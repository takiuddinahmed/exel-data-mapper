from cgi import FieldStorage
import os
from werkzeug.utils import secure_filename
from random import random
import openpyxl
dir_path = os.path.join(os.path.dirname(os.path.realpath(__file__)),"uploads")
if not (os.path.exists(dir_path)):
    os.mkdir(dir_path)

def upload_file(uploadedFile: FieldStorage)-> str:
    """
    Save the file and return the filename
    """
    filename = str(int(random() * 1000000)) + "-" + secure_filename(uploadedFile.name)  + ".xlsx";
    
    uploadedFile.save(os.path.join(dir_path, filename))
    return filename

def load_data(filename: str):
    print("loading....")
    path = os.path.join(dir_path,filename)
    print(path)
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    print(sheet_obj.max_column)
    print("End....")
    # for i in range(1, sheet_obj.max_column + 1):
    #     cell_obj = sheet_obj.cell(row = 1, column = i)
    #     print(cell_obj.value)
    arr = []
    for row in sheet_obj.values:
        row_arr = []
        for value in row:
            row_arr.append(value)
        arr.append(row_arr)
    return arr

def raw_data_to_dict(arr):
    dict_arr = []
    first_row = arr[0]
    other_row = arr[1:]
    for row in other_row:
        dict = {}
        i = 0
        for value in row:
            dict[first_row[i]] = value
            i=i+1
        dict_arr.append(dict)
    return dict_arr
